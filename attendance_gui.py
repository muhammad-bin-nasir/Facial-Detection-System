# attendance_gui.py

import sys
import os
import cv2
import face_recognition
import numpy as np
import pandas as pd
import datetime
import mediapipe as mp
from openpyxl import load_workbook
from PyQt6.QtWidgets import QApplication, QWidget, QLabel, QVBoxLayout, QPushButton, QTableWidget, QTableWidgetItem, QMessageBox, QInputDialog
from PyQt6.QtCore import QTimer, Qt
from PyQt6.QtGui import QImage, QPixmap

PASSWORD = "admin123"
KNOWN_FACES_DIR = "known_faces"
EXCEL_FILE = "attendance.xlsx"

known_encodings = []
known_names = []

# Load faces
for filename in os.listdir(KNOWN_FACES_DIR):
    img_path = os.path.join(KNOWN_FACES_DIR, filename)
    image = face_recognition.load_image_file(img_path)
    enc = face_recognition.face_encodings(image)
    if enc:
        known_encodings.append(enc[0])
        known_names.append(os.path.splitext(filename)[0])

mp_face_mesh = mp.solutions.face_mesh
face_mesh = mp_face_mesh.FaceMesh(refine_landmarks=True)

def is_real_face(frame):
    img_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    results = face_mesh.process(img_rgb)
    return results.multi_face_landmarks is not None

today = datetime.date.today().strftime("%Y-%m-%d")

if not os.path.exists(EXCEL_FILE):
    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
        pd.DataFrame(columns=["Name", "Time"]).to_excel(writer, sheet_name=today, index=False)

try:
    wb = load_workbook(EXCEL_FILE)
    df_today = pd.read_excel(EXCEL_FILE, sheet_name=today) if today in wb.sheetnames else pd.DataFrame(columns=["Name", "Time"])
    marked_today = set(df_today["Name"].values)
except Exception:
    marked_today = set()

class AttendanceApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Face Recognition Attendance System")
        self.setGeometry(100, 100, 800, 600)

        self.image_label = QLabel("Camera feed will appear here.")
        self.image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        self.status_label = QLabel("Status: Idle")
        self.attendance_table = QTableWidget(0, 2)
        self.attendance_table.setHorizontalHeaderLabels(["Name", "Time"])

        self.start_button = QPushButton("Start Camera")
        self.stop_button = QPushButton("Stop Camera")
        self.stop_button.setEnabled(False)

        layout = QVBoxLayout()
        layout.addWidget(self.image_label)
        layout.addWidget(self.status_label)
        layout.addWidget(self.attendance_table)
        layout.addWidget(self.start_button)
        layout.addWidget(self.stop_button)

        self.setLayout(layout)

        self.start_button.clicked.connect(self.start_camera)
        self.stop_button.clicked.connect(self.stop_camera)

        self.timer = QTimer()
        self.timer.timeout.connect(self.update_frame)
        self.cap = None
        self.add_button = QPushButton("Add New Face")
        self.remove_button = QPushButton("Remove Face")

        layout.addWidget(self.add_button)
        layout.addWidget(self.remove_button)

        self.add_button.clicked.connect(self.add_new_face)
        self.remove_button.clicked.connect(self.remove_face)

    def start_camera(self):
        self.cap = cv2.VideoCapture(0)
        self.timer.start(30)
        self.status_label.setText("Status: Camera started.")
        self.start_button.setEnabled(False)
        self.stop_button.setEnabled(True)

    def stop_camera(self):
        self.timer.stop()
        if self.cap:
            self.cap.release()
        self.image_label.clear()
        self.status_label.setText("Status: Camera stopped.")
        self.start_button.setEnabled(True)
        self.stop_button.setEnabled(False)
    def add_new_face(self):
        name, ok = QInputDialog.getText(self, "Add New Face", "Enter name:")
        if ok and name:
            cap = cv2.VideoCapture(0)
            ret, frame = cap.read()
            if ret:
                face_locations = face_recognition.face_locations(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
                if face_locations:
                    file_path = os.path.join(KNOWN_FACES_DIR, f"{name}.jpg")
                    cv2.imwrite(file_path, frame)
                    QMessageBox.information(self, "Success", f"Face of {name} added.")
                    self.update_known_faces()
                else:
                    QMessageBox.warning(self, "Warning", "No face detected. Try again.")
            cap.release()
        else:
            QMessageBox.warning(self, "Input Error", "No name provided.")

    def remove_face(self):
        if not known_names:
            QMessageBox.information(self, "Info", "No faces to remove.")
            return

        name, ok = QInputDialog.getItem(self, "Remove Face", "Select face to remove:", known_names, 0, False)
        if ok and name:
            file_path = os.path.join(KNOWN_FACES_DIR, f"{name}.jpg")
            if os.path.exists(file_path):
                os.remove(file_path)
                QMessageBox.information(self, "Removed", f"Face of {name} removed.")
                self.update_known_faces()
            else:
                QMessageBox.warning(self, "Error", "Face file not found.")

    def update_known_faces(self):
        global known_encodings, known_names
        known_encodings = []
        known_names = []

        for filename in os.listdir(KNOWN_FACES_DIR):
            img_path = os.path.join(KNOWN_FACES_DIR, filename)
            image = face_recognition.load_image_file(img_path)
            enc = face_recognition.face_encodings(image)
            if enc:
                known_encodings.append(enc[0])
                known_names.append(os.path.splitext(filename)[0])    

    def update_frame(self):
        ret, frame = self.cap.read()
        if not ret:
            return

        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        rgb_small = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

        face_locations = face_recognition.face_locations(rgb_small)
        face_encodings = face_recognition.face_encodings(rgb_small, face_locations)

        for face_encoding, face_location in zip(face_encodings, face_locations):
            matches = face_recognition.compare_faces(known_encodings, face_encoding)
            name = "Unknown"

            face_distances = face_recognition.face_distance(known_encodings, face_encoding)
            best_match = np.argmin(face_distances)

            if matches and matches[best_match]:
                name = known_names[best_match]

                if name not in marked_today and is_real_face(frame):
                    now = datetime.datetime.now()
                    time_str = now.strftime("%H:%M:%S")

                    df_new = pd.DataFrame([[name, time_str]], columns=["Name", "Time"])
                    try:
                        with pd.ExcelWriter(EXCEL_FILE, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
                            df_new.to_excel(writer, sheet_name=today, index=False, header=False, startrow=writer.sheets[today].max_row)
                    except KeyError:
                        with pd.ExcelWriter(EXCEL_FILE, mode="a", engine="openpyxl") as writer:
                            df_new.to_excel(writer, sheet_name=today, index=False)

                    marked_today.add(name)
                    row_pos = self.attendance_table.rowCount()
                    self.attendance_table.insertRow(row_pos)
                    self.attendance_table.setItem(row_pos, 0, QTableWidgetItem(name))
                    self.attendance_table.setItem(row_pos, 1, QTableWidgetItem(time_str))

                    self.status_label.setText(f"Status: {name} marked.")
                elif name in marked_today:
                    self.status_label.setText(f"Status: {name} already marked.")
                else:
                    self.status_label.setText("Status: Spoof attempt detected!")

            top, right, bottom, left = [v * 4 for v in face_location]
            cv2.rectangle(frame, (left, top), (right, bottom), (0,255,0) if name != "Unknown" else (0,0,255), 2)
            cv2.putText(frame, name, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255,255,255), 2)

        rgb_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb_image.shape
        bytes_per_line = ch * w
        qt_img = QImage(rgb_image.data, w, h, bytes_per_line, QImage.Format.Format_RGB888)
        self.image_label.setPixmap(QPixmap.fromImage(qt_img))

    def closeEvent(self, event):
        self.stop_camera()
        event.accept()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = AttendanceApp()
    window.show()
    sys.exit(app.exec())
